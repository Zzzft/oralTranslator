# main.py 批量处理口语翻译
import pandas as pd
import requests
from openpyxl import load_workbook
from tqdm import tqdm  # 进度条
from config import DEEPSEEK_API_KEY


URL = "https://api.siliconflow.cn/v1/chat/completions"

# 翻译规则
RULES = """
规则：
1.阅读上下文内容，根据原文提取人物和事件信息，分析出正确的句意。
2.请你删除口语表达中重复冗余的词汇，删除与句意无关的填充词。
例如：
- 嗯，那个，就是吧，我觉得这个事情呢，挺好的。（"嗯"，"那个"，"就是吧"，"呢" 是没有实质意义的填充词）
- 所以我刚才接着刚才说，所以他们基本都走。（刚才重复冗余）

3.请你将带有随意性和非正式性的口语词翻译成规范的书面语，将带有地域或代际特征的词语翻译成正式的书面表达。对于有多种含义的口语词汇，请根据句意替换为正确的表达。
例如：
- 她虚岁99岁走的。（“走的”即“去世”的口语说法）
- 这是09年的一个光是我们老爷们儿第二代和老娘照的，这是我们连三辈儿四辈儿都有的。（三辈儿、四辈儿即第三代、第四代人，老爷们儿指兄弟们，老娘指母亲）

4. 请你将指代模糊的人称代词替换为清晰的表达，根据句意推断出人称代词实际指向的人物。但如果上下文没有明确的指代对象，就保留原来的人称代词。
例如：
- 他告诉他，让他早点回家。（这里的三个“他”指代不明，需要根据上下文确定三个“他”分别指谁）
- 我弟妹说让我们把家腾干净了，我们要来住了（第二个“我们”是指说话的人，即弟妹他们）
- 他都所以所以在这种情况下，老三他们家提出来这个房子就应该是我们的。（这里的“我们”是指说话的人，即老三他们家）
- 从我们兄弟姐妹来说，受母亲的影响比较大，因为父亲那时候走的时候，他们有时候大部分时间在外地，刚回来也都挺忙的，正在中年，所以对父亲的印象不是特别深刻。(“他们”是指受父母影响的人，即“我们兄弟姐妹”)

5. 口语表达存在表达不准确或口误后自我修正的情况，对于这种前后表达存在冲突的情况，请保留修正后正确的结果。
例如：
- 我们哥五个哥六个觉得一定得这么办。（前半句“哥五个”为口误，后半句“哥六个”为修正，翻译后只需保留修正的后半句）

6. 当口语为了简化表达而省略一些句子成分（如主语、谓语或宾语）时，请你根据句意将这些成分补充完整。
例如：
- 母亲一直跟我在一起。（省略了（生活）在一起）
- 6个4个党员。（省略了6个（兄弟姐妹里））

7.当句子成分残缺（缺乏主谓宾或定状补）或句子结构复杂（多个从句或插入语）导致句子意思难以理解时，请你重新组织语言，调整语序，适当删减，使得句子易于理解。
例如：
- 母亲的影响尤其我们日后都退了休了又跟他在一起生活，影响越来越深刻了（成分残缺导致听者难以理解是母亲对谁的影响，应补充上母亲“对我们的”影响）。
- 南市住的话，她要是回一趟咸水沽的话，每次回去一趟起码得将近半天差不多，所以她也基本上不在家里，因为他们也是三班倒，那阵都是工人嘛。（“因为他们也是三班倒，那阵都是工人嘛”结构混乱，不易理解，实际后半句是对前半句的补充，是说“那时候大家都是工人”）
三班倒，所以这个时间基本上一个星期两个星期，所谓有个大大公休吧，那阵倒班，回家一次。（句子结构混乱，表述含糊，应该翻译为：她那时候“三班倒”，所以通常一到两个星期遇到大公休才能回家一次）
"""

def generate_prompt(oral, context):
    return f"""
你是一个老年人生活助手，需要将老年人的口语翻译成易于理解的简单语言，同时表达出完整的信息。请你按照以下规则翻译每一条语句。
{RULES}
    
你需要翻译的口语：
{oral}

口语所在上下文：
{context}

翻译完成后，请你核对翻译后的语句是否和原文意思相符，翻译后的句子是否通顺、前后句意是否相符，如果不符合原意或语句生硬，请修改润色。不要添加原文不存在的信息，也不要省略信息，确认无误后，请你将翻译后的结果告诉我。

仅输出指定口语的翻译结果，不要输出思考过程，不要输出解释。
"""
start_excel_row = 308    # 从 Excel 第 10 行开始
num_translations = 200   # 计划翻译 200 条

def process_excel(input_path):
    # 读取Excel文件
    df = pd.read_excel(input_path)
    
    # 确保存在结果列
    if '翻译结果' not in df.columns:
        df['翻译结果'] = ""
    
    # =========== 新增计算逻辑 ============
    # 计算 DataFrame 的起始索引（Excel行号转DataFrame索引）
    start_index = start_excel_row - 1
    if start_index < 0:
        start_index = 0

    # 计算实际可以处理的数量（防止溢出）
    end_index = start_index + num_translations
    actual_trans = min(num_translations, len(df) - start_index)

    # 初始化计数器
    processed_count = 0
    # ===================================

    with tqdm(total=actual_trans, desc="处理进度") as pbar:
        for index, row in df.iterrows():
            # 新增：只处理指定范围的行 [关键改动]
            if index < start_index:
                continue  # 跳过起始行之前的行

            # 新增：中断条件 [关键改动]
            if processed_count >= actual_trans:
                break

            # 原有逻辑（只处理空单元格）
            if pd.isnull(row['翻译结果']) or row['翻译结果'] == "":
                oral = str(row['原口语文本'])
                context = str(row['口语文本所在上下文'])
                
                translated = process_translation(oral, context)
                
                # 直接修改Excel
                book = load_workbook(input_path)
                sheet = book.active
                sheet.cell(row=index+2, column=5, value=translated)  # 行号修正映射
                book.save(input_path)
                book.close()

                # 更新计数器 
                processed_count += 1
                pbar.update(1)

            # 不论是否处理都要计数范围 [避免死循环]
            if index >= end_index:
                break

def process_translation(oral, context):
    prompt = generate_prompt(oral, context)
    
    payload = {
        "model": "deepseek-ai/DeepSeek-R1",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.2
    }
    
    headers = {
        "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
        "Content-Type": "application/json"
    }

    try:
        response = requests.post(URL, json=payload, headers=headers, timeout=180)
        response.raise_for_status()
        data = response.json()
        return data["choices"][0]["message"]["content"].strip()
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return "【翻译失败】"

if __name__ == "__main__":
    process_excel("data.xlsx")
    print("success!")
